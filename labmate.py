from tkinter import * 
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib
import pandas as pd

# REAGENT LOG
file = pathlib.Path("C:\\Users\\DELL\\Documents\\Labmate\\Reagent_log.xlsx")
if file.exists ():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']='User name'
    sheet['B1']='Item inuse'
    sheet['C1']='Quantity'
    sheet['D1']='Date'    
    sheet['E1']='Lot no.' 
    sheet['F1']='Date'
    sheet['G1']='Bench'
    sheet['H1']='Material'
    file.save("C:\\Users\\DELL\\Documents\\Labmate\\Reagent_log.xlsx")
 
# QC LOG    
file = pathlib.Path("C:\\Users\\DELL\\Documents\\Labmate\\QC_log.xlsx")
if file.exists ():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']= 'Qc issues'
    file.save("C:\\Users\\DELL\\Documents\\Labmate\\QC_log.xlsx")

    
# MACHINE LOG
file = pathlib.Path("C:\\Users\\DELL\\Documents\\Labmate\\Machine_log.xlsx")
if file.exists ():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']= 'Machine log'
    file.save("C:\\Users\\DELL\\Documents\\Labmate\\Machine_log.xlsx")
    

def submit():
    z=username.get()
    y=item_inuse.get()
    x=Quantity.get()
    w=date.get()
    v=lot_no.get()
    u=qc_issues.get()   

    print(z)
    print(y)
    print(x)
    print(w)
    print(v)
    print(u)
    
    file=openpyxl.load_workbook("C:\\Users\\DELL\\Documents\\Labmate\\QC_log.xlsx")
    sheet= file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=u)   
    file.save("C:\\Users\\DELL\\Documents\\Labmate\\QC_log.xlsx")
    
    file=openpyxl.load_workbook("C:\\Users\\DELL\\Documents\\Labmate\\Reagent_log.xlsx")
    sheet= file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=z)
    sheet.cell(column=2,row=sheet.max_row,value=y)
    sheet.cell(column=3,row=sheet.max_row,value=x)
    sheet.cell(column=4,row=sheet.max_row,value=w)
    sheet.cell(column=5,row=sheet.max_row,value=v)
    sheet.cell(column=6,row=sheet.max_row,value=u)

    
    if var1.get()==1:
        print('Endoserology')
        sheet.cell(column=7,row=sheet.max_row,value='Endoserology')
    if var1.get()==2:
        print('Chemistry')
        sheet.cell(column=7,row=sheet.max_row,value='Chemistry')
    if var1.get()==3:
        print('Haematology')
        sheet.cell(column=7,row=sheet.max_row,value='Haematology')
        
    if var2.get()==1:
        print('Reagent')
        sheet.cell(column=8,row=sheet.max_row,value='Reagent')
    if var2.get()==2:
        print('QC')
        sheet.cell(column=8,row=sheet.max_row,value='QC')
    if var2.get()==3:
        print('Calibrator')
        sheet.cell(column=8,row=sheet.max_row,value='Calibrator')
    if var2.get()==4:
        print('Consumable')
        sheet.cell(column=8,row=sheet.max_row,value='Consumables')  
        
    file.save("C:\\Users\\DELL\\Documents\\Labmate\\Reagent_log.xlsx")
    
#def clear_text(main):
    #main.entry.delete(0, 'end')
    #xlfile = pd.read_excel('Reagent_log.xlsx')
    #xlfile.to_csv('Reagent_log.csv', index=false)
    
#LABELS   
main = Tk()
main.title('Labmate')
main.geometry('1360x900')
main.config(highlightbackground='purple', highlightthickness=3)

frame1 = LabelFrame(main, text='Login:').pack(expand='yes', fill='both')   
Label(frame1, text='User:').place(x=30,y=30)
Label(frame1, text='Items put In Use:').place(x=30,y=70)
Label(frame1, text='Quantity:').place(x=30,y=110)
Label(frame1, text='Date:').place(x=450, y=30)
Label(frame1, text='Lot no:').place(x=450, y=70)
Label(frame1, text='QC Issues:').place(x=450, y=110)

#ENTRIES

username = Entry(frame1)
username.place(x=150, y=30)
item_inuse = Entry(frame1) 
item_inuse.place(x=150, y=70) 
Quantity = Entry(frame1)
Quantity.place(x=150, y=110) 
date = Entry(frame1)
date.place(x=570, y=30)
lot_no = Entry(frame1)
lot_no.place(x=570, y=70)
qc_issues = Entry(frame1,)
qc_issues.place(x=570, y=110)

             

frame2 = LabelFrame(main, text='Other details:').pack(expand='yes', fill='both')  
label_3= Label(frame2, text= 'Bench:', width=30, font=('bold',10))
label_3.place(relx=.03, rely=0.75)
var1=IntVar()
Radiobutton(frame2, text='Endoserology', variable=var1, value=1).place(relx=.25, rely=0.75)
Radiobutton(frame2, text='Chemistry', variable=var1, value=2).place(relx=.4, rely=0.75)
Radiobutton(frame2, text='Haematology', variable=var1, value=3).place(relx=.55, rely=0.75)

label_4=Label(frame2, text='Material:', width=30, font=('bold',10))
label_4.place(relx=.037, rely=0.63)
var2=IntVar()
Radiobutton(frame2, text='Reagent', padx=5, variable=var2, value=1).place(relx=.25, rely=0.63)
Radiobutton(frame2, text='QC', padx=5, variable=var2, value=2).place(relx=.4, rely=0.63)
Radiobutton(frame2, text='Calibrator', padx=5, variable=var2, value=3).place(relx=.49, rely=0.63)
Radiobutton(frame2, text='Consumable', padx=5, variable=var2, value=4).place(relx=.62, rely=0.63)

def clear():
    [widget.delete(0, tk.END) for widget in main.winfo_children() if isinstance(widget, tk.Entry)]


Button(frame2, text='Submit', command=submit).place(relx=.44, rely=.9,)
Button(frame2, text='Clear', command=clear).place(relx=.5, rely=.9)
'''
style = ttk.Style()
style.theme_use('default')
style.configure('Treeview',
            background='#D3D3D3',
            foreground= 'black',
            rowheight =20,
            fieldbackground='#D3D3D3')
style.map('Treeview') 
        #background=[('seleted', '#347083')])
tree_frame = Frame(main)
tree_frame.pack(pady = 10)
tree_scroll = Scrollbar(tree_frame)
tree_scroll.pack(side='right', fill =Y)

my_tree = ttk.Treeview(tree_frame, yscrollcommand= tree_scroll.set, selectmode='extended')
my_tree.pack() 

tree_scroll.config(command=my_tree.yview())

my_tree['columns']=('quantity', 'reagent')
my_tree.column('#0', width=0, stretch=NO)
my_tree.column('quantity', anchor=W, width=140)          
'''
top = Tk()


main.mainloop()
